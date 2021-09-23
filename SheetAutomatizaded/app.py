import openpyxl

# Creating a sheet(called book)
book = openpyxl.Workbook()

# How to visualize existents pages
print(book.sheetnames)

# How to create a page
book.create_sheet('Statistics')

# How to select a page
book_statistic = book['Statistics']

# Adding items in line
book_statistic.append(['Company', 'Stock', 'Price', 'Role', 'Group'])
book_statistic.append(['ITAUSA', 'ITSA4', 'R$10.45', 'Holding', 'Itaú UniBanco'])
book_statistic.append(['PETROBRAS', 'PETR4', 'R$25.85', 'Petróleo', 'Petrobras'])
book_statistic.append(['WEG', 'WEGE3', 'R$40.57', 'Máquinas Elétricas', 'WEG Eletric Corp'])
book_statistic.append(['CEMIG', 'CMIG4', 'R$14.05', 'Energia', 'Taesa S.A.'])

# Save
book.save('Sheet Stocks.xlsx')